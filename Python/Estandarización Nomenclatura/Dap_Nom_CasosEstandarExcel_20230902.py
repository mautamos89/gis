import tkinter
from tkinter import StringVar,CENTER
from tkinter import ttk,messagebox
from time import strftime

# Créditos y desarrollo

print("inicio de programa: ".upper() + strftime("%Y-%m-%d %H:%M:%S"))
print('título: programa para identificar componentes de la estandarización de Nomenclatura.'.upper())
print('licencia: Creative Commons. Condiciones de: atribución, no comercial, no derivada y compartir igual.'.upper())
print('requerimientos: sistema operativo windows'.upper())
print('grupo de desarrollo: nomenclatura domiciliaria, departamento administrativo de planeación - cali, 2023 ©.'.upper())
print('autor: mauricio tabares mosquera.'.upper())
print('profesión: geógrafo, especialista sig.'.upper())
print('versión: 1.0 | fecha: 2023-09-10.'.upper())


# Mensaje de bienvenida

messagebox.showinfo('Le damos la bienvenida!', """
La aplicación "Estandarización de Nomenclatura" ha sido
diseñada por el DAP con el objetivo de facilitar la
identificación de los componentes de la nomenclatura
abreviada a partir de la información oficial de la
Superintendencia de Notariado y Registro (SNR).\n
Si tiene alguna pregunta o desea verificar una imprecisión,
se sugiere hacer uso de la plataforma Ventanlla Única de
Registro (VUR) de la SNR. Además, se pueden consultar
plataformas de la Alcaldía de Cali como SAUL,Geovisor
IDESC y Geoportal Catastral.
""")

# Ventana de archivos

from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename
messagebox.showinfo('Selecciona el archivo!','Por favor, seleccione el archivo Excel a procesar.')
xfile=askopenfilename(filetypes=[('Archivos de Excel', ['*.xlsx','*.xls'])]) # Archivo a leer
print(f'ARCHIVO SELECCIONADO: {xfile}')
n_xfile = xfile[:-5] + f'_{strftime("%Y%m%d_%H%M%S")}.xlsx' # Copia para guardar
print(f'RUTA DEL ARCHIVO PROCESADO: {n_xfile}')
libro = load_workbook(f'{xfile}')
messagebox.showinfo('Archivo cargado con éxito',f'Archivo seleccionado: {xfile}\nHojas electrónicas del archivo: {libro.sheetnames}')
print(f'HOJAS DEL LIBRO: {libro.sheetnames}')

# GUI: Ventana principal

window = tkinter.Tk()
style=ttk.Style()
window.title('Estandarización de Nomenclatura | DAP, 2023 ©')

# Variables de búsqueda

usr_hoja_dat=StringVar() # Hoja con los datos
usr_col_dat=StringVar() # Columna con registros
usr_rango_dat=StringVar() # Rango máximo de datos
usr_col_res=StringVar() # La columna de resultados

# Recuadro principal: arriba

frame_upper=tkinter.Frame(window)
frame_upper.grid(column=0,row=0,sticky='nw')

# Recuadros secundarios:

# A: información sobre el desarrollo
fra_desa_app = tkinter.LabelFrame(frame_upper,text='Desarrollado por',font="TkDefaultFont 10")
fra_desa_app.grid(row=0,column=1,padx=12,pady=10,sticky='nw')
# B: datos de nomenclatura
fra_col_dat = tkinter.LabelFrame(frame_upper,text='Datos de la nomenclatura',font="TkDefaultFont 10")
fra_col_dat.grid(row=0,column=0,padx=12,pady=10,sticky='nw')

# Etiquetas:
from PIL import ImageTk
from urllib.request import urlopen
# A:

url='https://drive.google.com/uc?export=view&id=1yGQ2Xf4yyoXV545Kz8Pvp-7x1i1zf6X2'
u = urlopen(url)
raw_data = u.read()
u.close()
photo = ImageTk.PhotoImage(data=raw_data)
lab_desa_app=tkinter.Label(fra_desa_app,image=photo)
lab_desa_app.image = photo
lab_desa_app.grid(row=0,column=0,sticky='sw',padx=7,pady=4)
# B:
lab_cfg1,lab_cfg2,lab_cfg3 = "TkDefaultFont 10","#F5DEB3","#354A54"
lab_hoja_dat = tkinter.Label(fra_col_dat,text='Hoja electrónica:',font=lab_cfg1)
lab_hoja_dat.grid(row=0,column=0,sticky='nw')
# C:
lab_col_dat = tkinter.Label(fra_col_dat,text='Columna con registros:',font=lab_cfg1)
lab_col_dat.grid(row=1,column=0,sticky='nw')
# D:
lab_rango_dat = tkinter.Label(fra_col_dat,text='Celda máxima de datos:',font=lab_cfg1)
lab_rango_dat.grid(row=2,column=0,sticky='nw')
# E:
lab_alma_res = tkinter.Label(fra_col_dat,text='Columna para resultados:',font=lab_cfg1)
lab_alma_res.grid(row=3,column=0,sticky='nw')

# Entradas de usuario:

# A:
ent_hoja_dat = tkinter.Entry(fra_col_dat,width=15,textvariable=usr_hoja_dat)
ent_hoja_dat.grid(row=0,column=1)
# B:
ent_col_dat = tkinter.Entry(fra_col_dat,width=15,textvariable=usr_col_dat)
ent_col_dat.grid(row=1,column=1)
# C:
ent_rango_dat = tkinter.Entry(fra_col_dat,width=15,textvariable=usr_rango_dat)
ent_rango_dat.grid(row=2,column=1)
# D:
ent_alma_res = tkinter.Entry(fra_col_dat,width=15,textvariable=usr_col_res)
ent_alma_res.grid(row=3,column=1)

# Configuración de estilo de widgets

# A:
for widget in fra_desa_app.winfo_children(): widget.grid_configure(padx=5,pady=8.5)
# B:
for widget in fra_col_dat.winfo_children(): widget.grid_configure(padx=10,pady=7)

# Recuadro principal: abajo

frame_down=tkinter.Frame(window)
frame_down.grid(column=0,row=1)

# Recuadro # 2: botones interactivos

fra_usr_bot = tkinter.LabelFrame(frame_down,text='Botones interactivos',font="TkDefaultFont 10")
fra_usr_bot.grid(row=0,column=0,sticky='nw',padx=20,pady=10)

# Función: abecedario a números

d_alfa_num = {
    1:'a', 2:'b', 3:'c', 4:'d',
    5:'e', 6:'f', 7:'g', 8:'h',
    9:'i', 10:'j', 11:'k', 12:'l',
    13:'m', 14:'n', 15:'o', 16:'p',
    17:'q', 18:'r', 19:'s', 20:'t',
    21:'u', 22:'v', 23:'w', 24:'x',
    25:'y', 26:'z', 27:'aa', 28:'ab',
    29:'ac', 30:'ad', 31:'ae', 32:'af'
}
def x(a):
    if a in d_alfa_num.keys():
        #print(d_alfa_num[a].upper())
        return d_alfa_num[a].upper()
    if a in d_alfa_num.values():
        for key,val in d_alfa_num.items():
            if val == a:
                #print(key)
                return key
    else:
      return ''

# Función: estandarizar e identificar los casos de nomenclatura

def nom_casos():
    try:
        if usr_hoja_dat.get() and usr_col_dat.get() and usr_col_res.get() and usr_rango_dat.get():
            print("INICIO DEL PROCESO: ".upper() + strftime("%Y-%m-%d %H:%M:%S"))
            # Variables del ususario.
            e_hoja=usr_hoja_dat.get().strip()
            e_rcolumn=x(usr_col_dat.get().lower().strip())
            e_wcolumn=x(usr_col_res.get().lower().strip())
            e_max_data_row=int(usr_rango_dat.get().strip())
            # Variables de la función.
            hoja = libro[f'{e_hoja}'] # Hoja activa para trabajar
            rcolumn=e_rcolumn # Columna para leer datos (A=0,B=1,etc)
            wcolumn=e_wcolumn # Columna para escribir resultados
            start_col=wcolumn
            max_data_row=e_max_data_row # Rango máximo de filas en archivo
            """ hoja = libro["ph_c3_c5_c6"] # Hoja activa para trabajar
            rcolumn=12 # Columna para leer datos (A=0,B=1,etc)
            wcolumn=17 # Columna para escribir resultados
            start_col=wcolumn
            max_data_row=46231 # Rango máximo de filas en archivo """
            print(f'HOJA EXCEL: {e_hoja}\nCOLUMNA CON DATOS: {usr_col_dat.get().upper().strip()}\nRANGO DE DATOS: {e_max_data_row}\nCOLUMNA DE RESULTADOS: {usr_col_res.get().upper().strip()}')
            messagebox.showinfo('Párametros ingresados!', f'Hoja excel: {e_hoja}\nColumna con datos: {usr_col_dat.get().upper().strip()}\nRango de datos: {e_max_data_row}\nColumna de resultados: {usr_col_res.get().upper().strip()}')
            # Primer componente: columna + 0

            hoja.cell(1,wcolumn).value='COMP. 1'
            print('PROCESANDO COMPONENTE 1')
            for row in hoja.iter_cols(min_row=2, min_col=rcolumn, max_row=max_data_row, max_col=rcolumn):
                for cell in row:

                    # No impactar la tabla

                    if cell.value == None:
                        continue

                    # Impactar la tabla
                    
                    elif 'manzana' in cell.value.lower() or 'mz' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='MZ'
                    elif 'sector' in cell.value.lower() or 'sec' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='SEC'

            # Segundo componente: columna + 1

            wcolumn +=1
            hoja.cell(1,wcolumn).value='COMP. 2'
            print('PROCESANDO COMPONENTE 2')
            for row in hoja.iter_cols(min_row=2, min_col=rcolumn, max_row=max_data_row, max_col=rcolumn):
                for cell in row:

                    # No impactar la tabla

                    if cell.value == None:
                        continue

                    # Impactar la tabla

                    elif 'agrupacion' in cell.value.lower() or 'agrupación' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='AGP'

            # Tercer componente: columna + 1

            wcolumn +=1
            hoja.cell(1,wcolumn).value='COMP. 3'
            print('PROCESANDO COMPONENTE 3')
            for row in hoja.iter_cols(min_row=2, min_col=rcolumn, max_row=max_data_row, max_col=rcolumn):
                for cell in row:

                    # No impactar la tabla

                    if cell.value == None:
                        continue

                    # Impactar la tabla

                    elif 'edificio' in cell.value.lower() or 'edf' in cell.value.lower() or 'edif' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='ED'
                    elif 'bloque' in cell.value.lower() or 'blq' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='BLQ'
                    elif 'torre' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='TO'
                    elif 'interior' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='INT'

            # Cuarto componente: columna + 1

            wcolumn +=1
            hoja.cell(1,wcolumn).value='COMP. 4' # Se llega hasta el componente Sótano 4 como excepción
            print('PROCESANDO COMPONENTE 4')
            for row in hoja.iter_cols(min_row=2, min_col=rcolumn, max_row=max_data_row, max_col=rcolumn):
                for cell in row:
                    
                    # No impactar la tabla

                    if cell.value == None:
                        continue
                    elif 'piso' in cell.value.lower() and not ('apartamento' in cell.value.lower() or ('apto') in cell.value.lower() or 'penthouse' in cell.value.lower() or 'pent house' in cell.value.lower() or 'pen house' in cell.value.lower() or 'sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='P'
                    elif 'garaje' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()):
                        #print(f'semi: {cell.value.lower()}') # Verificar las celdas que toma la condición
                        continue
                    elif 'garaje' in cell.value.lower() and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()):
                        continue
                    elif 'parqueadero' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()):
                        continue
                    elif 'parqueadero' in cell.value.lower() and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()):
                        continue
                    elif 'oficina' in cell.value.lower() and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()):
                        continue
                    elif ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()):
                        continue
                    elif ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()):
                        continue
                    elif 'local' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()):
                        continue
                    elif 'local' in cell.value.lower() and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()):
                        continue
                    elif 'bodega' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()):
                        continue
                    elif 'bodega' in cell.value.lower() and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()):
                        continue
                    elif 'disponible' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()):
                        continue
                    elif 'disponible' in cell.value.lower() and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()):
                        continue
                    elif 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        continue
                    elif 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        continue
                    
                    # Impactar la tabla

                    elif 'semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='SS'
                    elif 'sotano' in cell.value.lower() or 'sótano' in cell.value.lower() or 'sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='SO'
                    elif 'mezzanine' in cell.value.lower() or 'mezanine' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='MN'
                    elif 'nivel' in cell.value.lower() and not ('apartamento' in cell.value.lower() or ('apto') in cell.value.lower() or 'penthouse' in cell.value.lower() or 'pent house' in cell.value.lower() or 'pen house' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='NIV'
                    #iba parqueo

            # Quinto componente: columna + 1

            wcolumn +=1
            hoja.cell(1,wcolumn).value='COMP. 5'
            print('PROCESANDO COMPONENTE 5')
            for row in hoja.iter_cols(min_row=2, min_col=rcolumn, max_row=max_data_row, max_col=rcolumn):
                for cell in row:
                    
                    # No impactar la tabla

                    if cell.value == None:
                        continue
                    
                    # Impactar la tabla

                    elif 'almacen' in cell.value.lower() or 'almacén' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='ALM'
                    elif 'penthouse' in cell.value.lower() or 'pent house' in cell.value.lower() or 'pen house' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PNH'
                    elif 'apto' in cell.value.lower() or 'apartamento' in cell.value.lower() or 'departamento' in cell.value.lower() or 'depto' in cell.value.lower() or 'dpto' in cell.value.lower() and not ('penthouse' in cell.value.lower() or ('pent house') in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='AP'
                    elif 'biblioteca' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='BIBL'
                    #iba bodega
                    elif 'burbuja' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='BURB'
                    elif 'cajero' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='CAJ'
                    #iba casa
                    elif 'celula' in cell.value.lower() or 'célula' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='CEL'
                    elif 'consultorio' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='CS'
                    #iba depósito
                    #iba disponible
                    elif 'garaje' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GADPSS'
                    elif 'garaje' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GADPS'
                    elif 'garaje' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GADP'
                    elif 'garaje' in cell.value.lower() and 'doble' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GADDP'
                    elif 'garaje' in cell.value.lower() and 'doble' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GADSS'
                    elif 'garaje' in cell.value.lower() and 'doble' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GADS'
                    elif 'garaje' in cell.value.lower() and 'doble' in cell.value.lower():# and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GAD'
                    elif 'garaje' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GASS'
                    elif 'garaje' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GAS'
                    elif 'garaje' in cell.value.lower() and 'oficina' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GAOF'
                    elif 'garaje' in cell.value.lower() or 'garaje' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='GA'
                    elif 'isla' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='IS'
                    elif 'kiosko' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='KCO'
                    elif 'laboratorio' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='LAB'
                    #iba local
                    elif 'modulo' in cell.value.lower() or 'módulo' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='MD'
                    elif 'oficina' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='OFS'
                    elif 'oficina' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='OF'
                    elif 'parqueadero' in cell.value.lower() and 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQCUSS'
                    elif 'parqueadero' in cell.value.lower() and 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQCUS'
                    elif 'parqueadero' in cell.value.lower() and 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQCU'
                    elif 'parqueadero' in cell.value.lower() and 'doble' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQDDP'
                    elif 'parqueadero' in cell.value.lower() and 'doble' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQDSS'
                    elif 'parqueadero' in cell.value.lower() and 'doble' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQDS'
                    elif 'parqueadero' in cell.value.lower() and 'doble' in cell.value.lower():# and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQD'
                    elif 'parqueadero' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQDPSS'
                    elif 'parqueadero' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not ('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQDPS'
                    elif 'parqueadero' in cell.value.lower() and ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQDP'
                    elif 'parqueadero' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQSS'
                    elif 'parqueadero' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQS'
                    #iba parqueadero
                    elif 'local' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='LCSS'
                    elif 'local' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='LCS'
                    elif 'local' in cell.value.lower() and ('mezzanine' in cell.value.lower() or 'mezanine' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='LCM'
                    elif 'local' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='LC'
                    elif 'parqueadero' in cell.value.lower() or 'estacionamiento' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PQ'
                    elif 'parqueo' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='PR'
                    #iba bodega
                    #iba local
                    elif 'casa' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='CA'
                    #iba cuarto útil
                    elif 'show' in cell.value.lower() and 'room' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='SHR'

            # Sexto componente: columna + 1

            wcolumn +=1
            hoja.cell(1,wcolumn).value='COMP. 6' # Se deja porque existe individual y combinado: DPS, PQ - DP
            print('PROCESANDO COMPONENTE 6')
            for row in hoja.iter_cols(min_row=2, min_col=rcolumn, max_row=max_data_row, max_col=rcolumn):
                for cell in row:

                    # No impactar la tabla

                    if cell.value == None:
                        continue
                    elif ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('parqueadero' in cell.value.lower() or 'garaje' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        continue
                    elif 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and ('parqueadero' in cell.value.lower() or 'garaje' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        continue

                    # Impactar la tabla

                    elif 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='CUSS'
                    elif 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='CUS'
                    elif 'cuarto' in cell.value.lower() and ('util' in cell.value.lower() or 'útil' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='CU'
                    elif ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='DPSS'
                    elif ('deposito' in cell.value.lower() or 'depósito' in cell.value.lower()) and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='DPS'
                    elif 'deposito' in cell.value.lower() or 'depósito' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='DP'
                    elif 'bodega' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='BDGASS'
                    elif 'bodega' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='BDGAS'
                    elif 'bodega' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='BDGA'
                    elif 'disponible' in cell.value.lower() and ('semisotano' in cell.value.lower() or 'semisótano' in cell.value.lower()) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='DISPSS'
                    elif 'disponible' in cell.value.lower() and ('sotano' in cell.value.lower() or 'sótano' in cell.value.lower()) and not (('sotano 1' in cell.value.lower() or 'sótano 1' in cell.value.lower() or 'sotano 2' in cell.value.lower() or 'sótano 2' in cell.value.lower() or 'sotano 3' in cell.value.lower() or 'sótano 3' in cell.value.lower() or 'sotano 4' in cell.value.lower() or 'sótano 4' in cell.value.lower())) and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='DISPS'
                    elif 'disponible' in cell.value.lower() and hoja.cell(cell.row,wcolumn).value == None:
                        hoja.cell(cell.row,wcolumn).value='DISP'
                    #Podría estar Mezzanine??
    
            # Calculando fórmula

            end_col = wcolumn
            wcolumn+=1
            #print(f'{wcolumn}-{x(wcolumn)}')
            #print(f'{x(wcolumn)}-{(start_col)}')
            hoja.cell(1,wcolumn).value='CASOS'
            print('CALCULANDO FÓRMULA PARA IDENTIFICAR LOS CASOS')
            for row in range(2, max_data_row+1):
                hoja[f"{x(wcolumn)}{row}"] = f'=_xlfn.TEXTJOIN("|",1,{x(start_col)}{row}:{x(end_col)}{row})' # Prefijo '_xlfn.' para usar fórmulas más nuevas de excel

            # Almacenando el archivo

            print(f'GUARDANDO ARCHIVO EN: {n_xfile}')
            libro.save(n_xfile)
            print("FINALIZACIÓN DEL PROCESO: ".upper() + strftime("%Y-%m-%d %H:%M:%S"))
    except:
        messagebox.showwarning('Problemas!', 'No se pudo procesar el archivo!\nNota: guarda el archivo con la extensión ".xls", luego en ".xlsx, luego intentas de nuevo"')

    else:
        messagebox.showinfo('Proceso terminado con éxito!', f'Resultado almacenado en: \n{n_xfile}')
# Botón procesar los casos

bot_repor=tkinter.Button(fra_usr_bot,text='construir patrón nomenclatura'.title(),command=nom_casos,anchor='center',justify=CENTER,width=30,height=1,font="bahnschrift 11",fg="white",bg='#2a9da1')
bot_repor.grid(row=0,column=0,sticky='nw',padx=10,pady=10)

# Iniciar interfaz gráfica

window.resizable(False,False)
window.mainloop()